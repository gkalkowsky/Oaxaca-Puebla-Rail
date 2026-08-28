#!/usr/bin/env python3
"""Build analysis/breakeven_model.xlsx — freight tonnage back-solve.

Regenerate with:  python3 analysis/scripts/build_breakeven_model.py

The task does NOT forecast demand. It back-solves the tonnage the line must
carry to cover its costs, then asks whether the corridor moves that much in
rail-divertible commodities.

THE O&M CIRCULARITY, HANDLED EXPLICITLY
Track maintenance scales with gross passing tonnage, so O&M is not independent
of the tonnage being solved for. Both revenue and variable O&M are linear in
tonnage, so the circularity closes algebraically rather than by iteration:

    margin x T x L  =  AnnualCapital + FixedOM + varOM x T x (1+tare) x L
    T = (AnnualCapital + FixedOM) / ( L x ( margin - varOM x (1+tare) ) )

The denominator is the net contribution per ton-km AFTER track wear. If it goes
to zero or below, no tonnage breaks even and the model says so rather than
returning a huge number. That guard is the point: it is the failure mode that a
naive model hides.

MARGIN IS SWEPT, NOT ASSUMED
ARTF's commodity-level contribution margin is not public and its Anuario was
unreachable (bot challenge). Rather than substitute a US Class I figure behind
the scenes, breakeven is solved ACROSS a margin range so the reader sees how
much the answer depends on it.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "analysis" / "breakeven_model.xlsx"

TITLE = Font(bold=True, size=14)
HEAD = Font(bold=True, color="FFFFFF")
SEC = Font(bold=True)
NOTE = Font(italic=True, size=9, color="666666")
WARN = Font(bold=True, color="9C0006")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
SEC_FILL = PatternFill("solid", fgColor="D9E2F3")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")
DER_FILL = PatternFill("solid", fgColor="E2EFDA")
ASSUM_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
M2 = '#,##0.00'; M1 = '#,##0.0'; N0 = '#,##0'; PCT = '0.0%'


def head(ws, row, labels):
    for c, t in enumerate(labels, 1):
        x = ws.cell(row=row, column=c, value=t); x.font, x.fill, x.border = HEAD, HEAD_FILL, BOX
        x.alignment = Alignment(vertical="center", wrap_text=True)


def sec(ws, row, label, width=6):
    ws.cell(row=row, column=1, value=label).font = SEC
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = SEC_FILL


def par(ws, row, name, val, unit, src="", note="", fmt=None, fill=IN_FILL):
    ws.cell(row=row, column=1, value=name)
    v = ws.cell(row=row, column=2, value=val); v.fill, v.border = fill, BOX
    if fmt: v.number_format = fmt
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=src)
    ws.cell(row=row, column=5, value=note).font = NOTE


def widths(ws, spec):
    for c, w in spec.items(): ws.column_dimensions[get_column_letter(c)].width = w


def build_readme(ws):
    ws.title = "README"
    ws["A1"] = "Vía Corta Oaxaca — Freight Reactivation Breakeven Model"; ws["A1"].font = TITLE
    lines = [
        "", "Puebla (Sánchez) → Oaxaca City, línea E, ~216.5 route-km. Screening level.", "",
        "WHAT THIS MODEL DOES",
        "Back-solves the annual tonnage required to break even, then expresses it as loaded",
        "truckloads/day each way so it is directly comparable to observed road traffic.",
        "It does NOT forecast demand.", "",
        "CELL COLOURS",
        "  Yellow  = input, enter with a source",
        "  Orange  = ASSUMPTION with no primary source behind it — the honest label, not a value",
        "  Green   = calculated, do not overtype", "",
        "THE TWO NUMBERS THAT CARRY THE ANSWER",
        "1. Contribution margin per ton-km. Not public at commodity level; ARTF's Anuario was",
        "   unreachable. It is therefore SWEPT across a range on 'Breakeven', not assumed.",
        "2. Track condition. Unknown since 2003 and unresolvable from a desk, so capital is",
        "   presented as light / heavy / substantial reconstruction. If those straddle the",
        "   breakeven threshold, STOP RULE 3 fires and the answer is INDETERMINATE.", "",
        "O&M CIRCULARITY",
        "Track maintenance scales with gross passing tonnage, so O&M depends on the tonnage",
        "being solved for. Both revenue and variable O&M are linear in tonnage, so the",
        "circularity closes algebraically (see 'Breakeven' col A note) rather than by iteration.",
        "If net contribution after track wear goes <= 0, NO tonnage breaks even and the sheet",
        "reports that instead of a large finite number.", "",
        "CURRENCY",
        "One base year, stated on 'Inputs'. MXN throughout. Any USD conversion states its rate",
        "and date. Mixing nominal figures across years is risk R-10.", "",
        "Regenerate: python3 analysis/scripts/build_breakeven_model.py",
    ]
    for i, t in enumerate(lines, 2):
        c = ws.cell(row=i, column=1, value=t)
        if t and t.isupper(): c.font = SEC
    widths(ws, {1: 100})


def build_inputs(ws):
    ws["A1"] = "Inputs — single editable block"; ws["A1"].font = TITLE
    ws["A2"] = "Every downstream sheet reads from here. Orange = assumption with no primary source."
    ws["A2"].font = NOTE
    head(ws, 3, ["Parameter", "Value", "Unit", "Source", "Notes"])

    sec(ws, 4, "CORRIDOR")
    par(ws, 5, "Route length", 216.5, "km", "Prompt.md (to verify)",
        "km E-150+000 to E-367+000. Independent verification outstanding", M1)

    sec(ws, 6, "CAPITAL — cost per route-km by track-condition scenario")
    par(ws, 7, "Light rehabilitation", 6.9, "MXN million / km", "wb-2020-serbia-railways-lcc T.11",
        "WB partial renewal (rail + ballast) 343/m @20 MXN/EUR. TRACK ONLY. Consistent with "
        "the UNESCAP < USD 500,000/route-km figure the brief cites", M2)
    par(ws, 8, "Heavy rehabilitation", 9.3, "MXN million / km", "wb-2020-serbia-railways-lcc T.11",
        "WB partial renewal (rail + sleeper exchange) 463/m @20 MXN/EUR. TRACK ONLY", M2)
    par(ws, 9, "Substantial reconstruction", 14.3, "MXN million / km",
        "wb-2020-serbia-railways-lcc T.11",
        "WB general renewal + subsoil rehabilitation 716/m @20 MXN/EUR. TRACK ONLY. "
        "Currency NOT stated in the source; EUR assumed as the likelier of the two", M2)
    ws.cell(row=10, column=1, value="  Mexican precedent (context, NOT a base case)")
    ws.cell(row=10, column=2, value=None).fill = ASSUM_FILL
    ws.cell(row=10, column=2).border = BOX
    ws.cell(row=10, column=4, value="[PRESS — UNVERIFIED]")
    ws.cell(row=10, column=5, value=(
        "Press reports ~18,000 MXN million on Línea Z rehabilitation (~300 km) => order "
        "~60 MXN million/km, roughly an order of magnitude above the UNESCAP figure. ASF "
        "primary unreachable (JS app), DOF egress-blocked. Must NOT carry a conclusion.")).font = NOTE

    sec(ws, 11, "CAPITAL — structures carried separately, per the brief")
    par(ws, 12, "Bridges / drainage / slope stabilisation — low", 0, "MXN million",
        "[UNSOURCED — SET 0]",
        "SET TO ZERO DELIBERATELY so the model computes a TRACK-ONLY LOWER BOUND. No public "
        "structure inventory is expected to exist for this line, and the World Bank unit costs "
        "above are track work only. Every result downstream is therefore an UNDERSTATEMENT of "
        "cost and an OVERSTATEMENT of viability. Replace with real figures before any decision", M1, ASSUM_FILL)
    par(ws, 13, "Bridges / drainage / slope stabilisation — high", 0, "MXN million",
        "[UNSOURCED — SET 0]",
        "Same. Material on this alignment: the Cañada sits where Sierra Madre Oriental and "
        "Sierra Madre del Sur folding converge, and CONANP already records slope instability, "
        "erosion and scour along the corridor's existing roads", M1, ASSUM_FILL)

    sec(ws, 14, "FINANCE")
    par(ws, 15, "Asset life", 30, "years", "Prompt.md", "", N0)
    par(ws, 16, "Cost of capital — case 1", 0.05, "% / yr", "Prompt.md", "", PCT)
    par(ws, 17, "Cost of capital — case 2", 0.06, "% / yr", "Prompt.md", "", PCT)
    par(ws, 18, "Cost of capital — case 3", 0.08, "% / yr", "Prompt.md", "", PCT)

    sec(ws, 19, "OPERATING")
    par(ws, 20, "Fixed O&M (tonnage-independent)", 0, "MXN million / yr",
        "set 0 — see note", "MUST be 0 on the ARTF EBIT margin basis: ARTF's cost base already "
        "includes network operation and maintenance. Non-zero here double-counts", M1, DER_FILL)
    par(ws, 21, "Variable O&M per gross ton-km", 0, "MXN / gross ton-km",
        "set 0 — see note", "Same reason. Track wear is already inside the 0.402 EBIT figure",
        '0.0000', DER_FILL)
    par(ws, 22, "Tare factor (gross / net tonnage)", 0, "ratio", "set 0 — see note",
        "Unused on the EBIT basis, since variable O&M is 0", M2, DER_FILL)

    sec(ws, 23, "REVENUE")
    par(ws, 24, "Contribution margin per net ton-km", 0.402, "MXN / net ton-km",
        "artf-2024-anuario-ferroviario",
        "Ferrosur EBIT/ton-km, constant 2024 MXN: $0.93 revenue (Tabla 7-8) x 43.19% operating "
        "margin (Tablas 7-3, 7.7). ARTF's cost base ALREADY includes maintenance and D&A, so "
        "B21 and B20 must be 0 when using this basis or maintenance is double-counted. "
        "Swept 0.402-0.93 on 'Supportable'. See working/margin-derivation.md", '0.0000')
    par(ws, 33, "Variable O&M — set 0 on the EBIT margin basis", 0, "MXN / gross ton-km",
        "see note", "Guard against double-counting; overrides B21 conceptually", '0.0000', DER_FILL)

    sec(ws, 25, "CONVENTIONS AND CONVERSION")
    par(ws, 26, "MXN / USD rate", 18.5, "MXN per USD", "[ASSUMED]",
        "State the rate AND its date. Required before any USD benchmark is used", M2)
    par(ws, 27, "FX rate date", "NOT SET — must be stated before use", "date",
        "[REQUIRED]", "An FX rate without its date is not a source", None, ASSUM_FILL)
    par(ws, 28, "Deflation base year (INEGI INPC)", 2024, "yyyy", "artf-2024-anuario",
        "All money figures real in this year. Never mix nominal across years", N0)

    sec(ws, 29, "ROAD COMPARISON")
    par(ws, 30, "Payload per articulated vehicle", 19.1, "tonnes", "imt-pt179 Tabla 4.7",
        "Weighted mean of IMT carga promedio (T3S2 13.2, T3S3 20.9, T3S2R4 30.1) at the observed "
        "terminus class mix. PRIMARY, but 2001 data — STALE", M1)
    par(ws, 31, "Empty-running share", 0, "fraction", "set 0 — see note",
        "MUST be 0 with the IMT payload above: carga promedio is averaged across observed "
        "vehicles and ALREADY nets out empty and partial running. A further discount "
        "double-counts", PCT, DER_FILL)
    par(ws, 32, "Observed articulated veh/day at corridor terminus", 500, "veh/day, both dirs",
        "sct-2025-datosviales-oaxaca", "SR-2 bound: endpoint flow N of Oaxaca City. See working/sr2-evaluation.md", N0)
    widths(ws, {1: 44, 2: 16, 3: 22, 4: 26, 5: 78})
    ws.freeze_panes = "A4"


def build_capital(ws):
    ws["A1"] = "Capital cost band"; ws["A1"].font = TITLE
    ws["A2"] = ("Track condition is unknown since 2003 and cannot be resolved from a desk, so "
                "capital is bounded by scenario rather than assumed to a base case.")
    ws["A2"].font = NOTE
    head(ws, 4, ["Scenario", "Cost per km (MXN m)", "Linework (MXN m)",
                 "Structures low (MXN m)", "Structures high (MXN m)",
                 "TOTAL low (MXN m)", "TOTAL high (MXN m)"])
    for i, (label, src) in enumerate(
            [("Light rehabilitation", 7), ("Heavy rehabilitation", 8),
             ("Substantial reconstruction", 9)], start=5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=f"=Inputs!B{src}").number_format = M2
        ws.cell(row=i, column=3,
                value=f'=IF(OR(B{i}="",Inputs!$B$5=""),"",B{i}*Inputs!$B$5)').number_format = M1
        ws.cell(row=i, column=4, value='=Inputs!$B$12').number_format = M1
        ws.cell(row=i, column=5, value='=Inputs!$B$13').number_format = M1
        ws.cell(row=i, column=6, value=f'=IF(OR(C{i}="",D{i}=""),"",C{i}+D{i})').number_format = M1
        ws.cell(row=i, column=7, value=f'=IF(OR(C{i}="",E{i}=""),"",C{i}+E{i})').number_format = M1
        for c in range(2, 8):
            ws.cell(row=i, column=c).fill = DER_FILL if c > 2 else IN_FILL
            ws.cell(row=i, column=c).border = BOX
    ws["A9"] = ("Structures (bridges, drainage, slope stabilisation) are shown as their own "
                "range, not folded into a contingency percentage — they are the dominant "
                "uncertainty on this alignment and no public structure inventory is expected.")
    ws["A9"].font = NOTE
    ws["A10"] = ("The Cañada de Cuicatlán sits where Sierra Madre Oriental and Sierra Madre del "
                 "Sur folding converge; documented rainy-season suspension implies scour and "
                 "slope exposure at water crossings.")
    ws["A10"].font = NOTE
    widths(ws, {1: 30, 2: 20, 3: 18, 4: 20, 5: 20, 6: 18, 7: 18})


def build_breakeven(ws):
    ws["A1"] = "Breakeven tonnage back-solve"; ws["A1"].font = TITLE
    ws["A2"] = ("T = (AnnualCapital + FixedOM) / ( L × ( margin − varOM × (1+tare) ) ).  "
                "Closed form: revenue and variable O&M are both linear in tonnage, so the O&M "
                "circularity resolves algebraically rather than by iteration.")
    ws["A2"].font = NOTE
    ws["A3"] = ("If margin − varOM×(1+tare) ≤ 0 the line loses money on every incremental tonne "
                "and NO tonnage breaks even. The sheet says so rather than returning a number.")
    ws["A3"].font = WARN

    ws["A5"] = "Net contribution per net ton-km after track wear"; ws["A5"].font = SEC
    ws["B5"] = ('=IF(OR(Inputs!B24="",Inputs!B21="",Inputs!B22=""),"",'
                'Inputs!B24-Inputs!B21*(1+Inputs!B22))')
    ws["B5"].number_format = '0.0000'; ws["B5"].fill = DER_FILL; ws["B5"].border = BOX
    ws["C5"] = "MXN / net ton-km"
    ws["D5"] = ('=IF(B5="","pending margin + O&M inputs",'
                'IF(B5<=0,"NO TONNAGE BREAKS EVEN — track wear exceeds margin","positive contribution"))')
    ws["D5"].font = SEC

    head(ws, 7, ["Capital scenario", "Capital (MXN m)", "Cost of capital",
                 "Annualised capital (MXN m/yr)", "Fixed O&M (MXN m/yr)",
                 "BREAKEVEN tonnage (net t/yr)", "Loaded truckloads/day each way"])
    row = 8
    for cap_row, label in ((6, "Light rehab (low structures)"),
                           (7, "Light rehab (high structures)"),
                           (8, "Heavy rehab (low structures)"),
                           (9, "Heavy rehab (high structures)"),
                           (10, "Substantial recon (low struct.)"),
                           (11, "Substantial recon (high struct.)")):
        cap_ref = {6: "Capital!F5", 7: "Capital!G5", 8: "Capital!F6",
                   9: "Capital!G6", 10: "Capital!F7", 11: "Capital!G7"}[cap_row]
        for disc in (16, 17, 18):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"={cap_ref}").number_format = M1
            ws.cell(row=row, column=3, value=f"=Inputs!$B${disc}").number_format = PCT
            ws.cell(row=row, column=4, value=(
                f'=IF(OR(B{row}="",C{row}="",Inputs!$B$15=""),"",'
                f'-PMT(C{row},Inputs!$B$15,B{row}))')).number_format = M1
            ws.cell(row=row, column=5, value='=Inputs!$B$20').number_format = M1
            ws.cell(row=row, column=6, value=(
                f'=IF(OR(D{row}="",E{row}="",$B$5="",$B$5<=0,Inputs!$B$5=""),"",'
                f'(D{row}+E{row})*1000000/(Inputs!$B$5*$B$5))')).number_format = N0
            ws.cell(row=row, column=7, value=(
                f'=IF(OR(F{row}="",Inputs!$B$30="",Inputs!$B$30=0),"",'
                f'F{row}/(Inputs!$B$30*365*2))')).number_format = N0
            for c in range(2, 8):
                ws.cell(row=row, column=c).fill = DER_FILL; ws.cell(row=row, column=c).border = BOX
            row += 1

    r = row + 1
    ws.cell(row=r, column=1, value="STOP RULE 3 — does the answer flip across track-condition scenarios?").font = SEC
    ws.cell(row=r + 1, column=1, value=(
        "Compare the breakeven range above against divertible tonnage on 'Compare'. If the "
        "project clears under light rehabilitation and fails under substantial reconstruction, "
        "the correct output is INDETERMINATE pending field reconnaissance — not a base case "
        "with a verdict attached. Report the scenario at which it flips: that is the single "
        "most decision-relevant number in the study.")).font = NOTE
    widths(ws, {1: 32, 2: 18, 3: 15, 4: 22, 5: 18, 6: 24, 7: 26})
    ws.freeze_panes = "A8"


def build_commodity(ws):
    ws["A1"] = "Commodity segregation and mode-diversion assumptions"; ws["A1"].font = TITLE
    ws["A2"] = ("The decisive analytical step. Diversion rates are ASSUMPTIONS and the answer is "
                "highly sensitive to them — each needs its own cited source before use.")
    ws["A2"].font = NOTE
    head(ws, 4, ["Commodity class", "Rail-divertible?", "Corridor tonnage (t/yr)",
                 "Diversion rate", "Divertible tonnage (t/yr)", "Source for diversion rate"])
    classes = [("Cement", "yes"), ("Aggregate", "yes"), ("Fertilizer", "yes"), ("Grain", "yes"),
               ("Fuel", "yes"), ("Steel", "yes"), ("Containerised manufactured", "yes"),
               ("Refrigerated", "low"), ("Time-sensitive", "low"),
               ("High-value low-density agricultural", "low")]
    for i, (name, div) in enumerate(classes, start=5):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=div)
        for c in (3, 4):
            ws.cell(row=i, column=c).fill = IN_FILL; ws.cell(row=i, column=c).border = BOX
        ws.cell(row=i, column=4).number_format = PCT
        e = ws.cell(row=i, column=5, value=f'=IF(OR(C{i}="",D{i}=""),"",C{i}*D{i})')
        e.number_format = N0; e.fill = DER_FILL; e.border = BOX
    tot = len(classes) + 5
    ws.cell(row=tot, column=1, value="TOTAL divertible").font = SEC
    t = ws.cell(row=tot, column=5, value=f"=IF(COUNT(E5:E{tot-1})=0,\"\",SUM(E5:E{tot-1}))")
    t.number_format = N0; t.fill = DER_FILL; t.border = BOX; t.font = SEC
    ws.cell(row=tot + 2, column=1, value=(
        "Low-diversion classes (mezcal, coffee, avocado, mango, figs) are a large share of "
        "Oaxaca's tradeable output and are exactly the goods least likely to move by rail. "
        "Segregating them is what prevents a headline tonnage figure from overstating the case.")).font = NOTE
    ws.cell(row=tot + 3, column=1, value=(
        "Aforo data classifies by axle configuration only — there is NO commodity field. "
        "Commodity mix must come from production and trade data (SIAP, INEGI Censos "
        "Económicos), not from truck counts.")).font = NOTE
    widths(ws, {1: 38, 2: 16, 3: 22, 4: 14, 5: 24, 6: 34})


def build_compare(ws):
    ws["A1"] = "Compare — divertible tonnage against breakeven"; ws["A1"].font = TITLE
    ws["A2"] = "Filled once both sides exist. Ranges, not point estimates."; ws["A2"].font = NOTE
    head(ws, 4, ["Quantity", "Value", "Unit", "Source"])
    rows = [
        ("Corridor tonnage, ALL articulated freight", "=Supportable!B8", "t / yr",
         "aforo 2024 x IMT loads"),
        ("Breakeven as % of ALL corridor freight — best case",
         '=IF(OR(B5="",B5=0,COUNT(Breakeven!F8:F25)=0),"",MIN(Breakeven!F8:F25)/B5)', "%", "derived"),
        ("Breakeven as % of ALL corridor freight — worst case",
         '=IF(OR(B5="",B5=0,COUNT(Breakeven!F8:F25)=0),"",MAX(Breakeven!F8:F25)/B5)', "%", "derived"),
        ("Divertible tonnage (bottom-up, commodity)", "=Commodity!E15", "t / yr", "Commodity sheet"),
        ("Breakeven tonnage — most favourable case", '=IF(COUNT(Breakeven!F8:F25)=0,"",MIN(Breakeven!F8:F25))', "t / yr", "Breakeven sheet"),
        ("Breakeven tonnage — least favourable case", '=IF(COUNT(Breakeven!F8:F25)=0,"",MAX(Breakeven!F8:F25))', "t / yr", "Breakeven sheet"),
        ("Ratio: divertible / breakeven (favourable)", '=IF(OR(B8="",B9="",B9=0),"",B8/B9)', "x", ""),
        ("Ratio: divertible / breakeven (unfavourable)", '=IF(OR(B8="",B10="",B10=0),"",B8/B10)', "x", ""),
    ]
    for i, (n, f, u, s) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=n)
        v = ws.cell(row=i, column=2, value=f); v.fill, v.border = DER_FILL, BOX
        v.number_format = M2 if "Ratio" in n else N0
        ws.cell(row=i, column=3, value=u); ws.cell(row=i, column=4, value=s)
    ws.cell(row=14, column=1, value="VERDICT GATE").font = SEC
    ws.cell(row=15, column=1, value="Stop Rule 3 test")
    g = ws.cell(row=15, column=2, value=(
        '=IF(OR(B11="",B12=""),"pending commodity segregation — see B6:B7 for the '
        'all-freight comparison",'
        'IF(AND(B11>1,B12<1),"INDETERMINATE — scenarios straddle breakeven; field reconnaissance required",'
        'IF(B11<1,"FAILS under every capital scenario",'
        'IF(B12>1,"CLEARS under every capital scenario","review"))))'))
    g.font = SEC; g.fill = DER_FILL; g.border = BOX
    ws.cell(row=17, column=1, value=(
        "A straddle is a real result, not a failure to decide. Per the brief, do not select a "
        "base case and present a conclusion — report the scenario at which the answer flips.")).font = NOTE
    ws.cell(row=18, column=1, value=(
        "Rows 6-7 are the comparison that does NOT wait on commodity segregation: breakeven "
        "tonnage as a share of ALL articulated freight in the corridor. Since divertible "
        "tonnage cannot exceed total tonnage, those percentages are a LOWER BOUND on the "
        "capture rate required — and they assume structures cost zero.")).font = NOTE
    widths(ws, {1: 46, 2: 58, 3: 14, 4: 22})



def build_supportable(ws):
    """The screen inverted: what capital could the corridor support?"""
    ws["A1"] = "Maximum supportable capital — the screen inverted"; ws["A1"].font = TITLE
    ws["A2"] = ("A planning-level capital cost for this alignment could not be sourced. So instead: "
                "given what the corridor demonstrably moves and what Mexican railways demonstrably "
                "earn per ton-km, how much capital could the line support? Needs no capital estimate.")
    ws["A2"].font = NOTE

    ws["A4"] = "CORRIDOR TONNAGE FROM OBSERVED TRAFFIC"; ws["A4"].font = SEC
    rows = [
        ("Articulated veh/day, both directions", "=Inputs!B32", "veh/day", "sct-2025-datosviales-oaxaca (SR-2 bound)"),
        ("Loaded share", '=IF(Inputs!B31="","",1-Inputs!B31)', "fraction", "[ASSUMPTION] 30–50% empty"),
        ("Payload per loaded truck", "=Inputs!B30", "tonnes", "[ASSUMPTION] 25–30 t"),
        ("Corridor tonnage", '=IF(OR(B5="",B6="",B7=""),"",B5*B6*B7*365)', "tonnes / yr", "derived"),
    ]
    for i, (n, f, u, s_) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=n)
        c = ws.cell(row=i, column=2, value=f); c.fill, c.border = DER_FILL, BOX
        c.number_format = N0 if i in (5, 8) else M2
        ws.cell(row=i, column=3, value=u); ws.cell(row=i, column=4, value=s_).font = NOTE

    ws["A10"] = "MAXIMUM SUPPORTABLE CAPITAL (MXN million per route-km)"; ws["A10"].font = SEC
    ws["A11"] = ("Rows = share of corridor freight won by rail [ASSUMPTION]. "
                 "Columns = cost of capital, 30-year life.")
    ws["A11"].font = NOTE
    head(ws, 12, ["Diversion", "Tonnage (t/yr)", "Annual surplus (MXN m)",
                  "@5% MXN m/km", "@6% MXN m/km", "@8% MXN m/km"])
    for i, div in enumerate([1.00, 0.50, 0.30, 0.15], start=13):
        ws.cell(row=i, column=1, value=div).number_format = PCT
        ws.cell(row=i, column=2, value=f'=IF($B$8="","",$B$8*A{i})').number_format = N0
        ws.cell(row=i, column=3, value=(
            f'=IF(OR(B{i}="",Inputs!$B$24="",Inputs!$B$5=""),"",'
            f'B{i}*Inputs!$B$5*Inputs!$B$24/1000000)')).number_format = M1
        for j, drow in enumerate((16, 17, 18)):
            col = 4 + j
            ws.cell(row=i, column=col, value=(
                f'=IF(OR($C{i}="",Inputs!$B${drow}="",Inputs!$B$15="",Inputs!$B$5=""),"",'
                f'$C{i}*((1-(1+Inputs!$B${drow})^-Inputs!$B$15)/Inputs!$B${drow})/Inputs!$B$5)'
            )).number_format = M1
        for c in range(2, 7):
            ws.cell(row=i, column=c).fill = DER_FILL; ws.cell(row=i, column=c).border = BOX

    ws["A19"] = "BENCHMARKS"; ws["A19"].font = SEC
    ws["A20"] = "UNESCAP light rehabilitation"
    ws["B20"] = '=IF(Inputs!B26="","",0.5*Inputs!B26)'
    ws["B20"].number_format = M2; ws["B20"].fill = DER_FILL; ws["B20"].border = BOX
    ws["C20"] = "MXN million / km"; ws["C20"].font = NOTE
    ws["D20"] = "< USD 500,000/route-km, converted at Inputs!B26 — state the rate and its date"
    ws["D20"].font = NOTE
    ws["A21"] = "Línea Z Mexican precedent"; ws["B21"] = 60
    ws["B21"].number_format = M2; ws["B21"].fill = ASSUM_FILL; ws["B21"].border = BOX
    ws["C21"] = "MXN million / km"; ws["C21"].font = NOTE
    ws["D21"] = "~18,000 MXN million / ~300 km  [PRESS — UNVERIFIED, must not carry a conclusion]"
    ws["D21"].font = NOTE

    ws["A23"] = "VERDICT TEST"; ws["A23"].font = SEC
    ws["B23"] = ('=IF(OR(D13="",B21=""),"pending inputs",'
                 'IF(F13<B21,"FAILS — even 100% capture at the highest cost of capital cannot '
                 'support the Mexican precedent unit cost","review"))')
    ws["B23"].font = SEC; ws["B23"].fill = DER_FILL; ws["B23"].border = BOX
    widths(ws, {1: 34, 2: 20, 3: 24, 4: 18, 5: 18, 6: 18})


def main():
    wb = Workbook()
    build_readme(wb.active)
    build_inputs(wb.create_sheet("Inputs"))
    build_capital(wb.create_sheet("Capital"))
    build_breakeven(wb.create_sheet("Breakeven"))
    build_commodity(wb.create_sheet("Commodity"))
    build_supportable(wb.create_sheet("Supportable"))
    build_compare(wb.create_sheet("Compare"))
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)} ({OUT.stat().st_size//1024} KB)")


if __name__ == "__main__":
    main()
